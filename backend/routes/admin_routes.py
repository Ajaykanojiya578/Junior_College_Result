# backend/routes/admin_routes.py

from flask import Blueprint, request, jsonify
from sqlalchemy.exc import IntegrityError

from app import db
from typing import Any, Dict, Optional, cast
from auth import generate_token
from models import (
    Teacher,
    Subject,
    Student,
    TeacherSubjectAllocation
)
from schemas import StudentSchema
from auth import token_required
from decorators import admin_required
from services.result_service import generate_results_for_division
from models import Result, Subject, Mark
from flask import send_file
from io import BytesIO

letter: Optional[Any] = None
canvas_module: Optional[Any] = None
try:
    import importlib
    _rl_pages = importlib.import_module('reportlab.lib.pagesizes')
    _rl_canvas = importlib.import_module('reportlab.pdfgen.canvas')
    letter = getattr(_rl_pages, 'letter', None)
    canvas_module = _rl_canvas
except Exception:
    letter = None
    canvas_module = None

from werkzeug.security import generate_password_hash


admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

student_schema = StudentSchema()


# ======================================================
# 1️⃣ Add Student
# ======================================================
@admin_bp.route("/students", methods=["POST"])
@token_required
@admin_required
def add_student(user_id=None, user_type=None):
    """
    Add a new student (admin only)
    """
    data = cast(Dict[str, Any], student_schema.load(request.json or {}))

    student = Student(**data)

    try:
        db.session.add(student)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return {"error": "Student already exists"}, 409

    return {"message": "Student added successfully"}, 201


# ======================================================
# 2️⃣ Get Students (by division)
# ======================================================
@admin_bp.route("/students", methods=["GET"])
@token_required
@admin_required
def list_students(user_id=None, user_type=None):
    """
    List students by division (admin only)
    """
    division = request.args.get("division")
    if not division:
        return {"error": "division is required"}, 400

    students = (
        Student.query
        .filter_by(division=division)
        .order_by(Student.roll_no)
        .all()
    )

    return jsonify([
        {
            "roll_no": s.roll_no,
            "name": s.name,
            "division": s.division,
            "optional_subject": s.optional_subject,
            "optional_subject_2": s.optional_subject_2
        }
        for s in students
    ]), 200


# ======================================================
# 3️⃣ Assign Teacher to Subject + Division
# ======================================================
@admin_bp.route("/allocations", methods=["POST"])
@token_required
@admin_required
def allocate_teacher(user_id=None, user_type=None):
    """
    Assign teacher to subject & division (admin only)
    """
    data: Dict[str, Any] = (request.json or {})
    teacher_id = data.get("teacher_id")
    subject_id = data.get("subject_id")
    division = data.get("division")

    if not teacher_id or not subject_id or not division:
        return {"error": "teacher_id, subject_id, division required"}, 400

    allocation = TeacherSubjectAllocation()
    allocation.teacher_id = teacher_id
    allocation.subject_id = subject_id
    allocation.division = division

    try:
        db.session.add(allocation)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return {"error": "Allocation already exists"}, 409

    return {"message": "Teacher allocated successfully"}, 201


@admin_bp.route('/allocations/<int:allocation_id>', methods=['DELETE'])
@token_required
@admin_required
def delete_allocation(allocation_id, user_id=None, user_type=None):
    """Delete a specific teacher-subject allocation (admin only)."""
    alloc = TeacherSubjectAllocation.query.get(allocation_id)
    if not alloc:
        return {"error": "Allocation not found"}, 404

    try:
        db.session.delete(alloc)
        db.session.commit()
        return {"message": "Allocation deleted"}, 200
    except Exception as ex:
        db.session.rollback()
        return {"error": "Failed to delete allocation", "details": str(ex)}, 500


# ======================================================
# 4️⃣ View Allocations
# ======================================================
@admin_bp.route("/allocations", methods=["GET"])
@token_required
@admin_required
def list_allocations(user_id=None, user_type=None):
    """
    List all teacher-subject allocations (admin only)
    """
    allocations = TeacherSubjectAllocation.query.all()

    result = []
    for a in allocations:
        teacher = Teacher.query.get(a.teacher_id)
        subj = Subject.query.get(a.subject_id)
        result.append({
            "allocation_id": a.allocation_id,
            "teacher_id": a.teacher_id,
            "teacher_name": teacher.name if teacher else None,
            "subject_id": a.subject_id,
            "subject_code": subj.subject_code if subj else None,
            "subject_name": subj.subject_name if subj else None,
            "division": a.division
        })

    return jsonify(result), 200


# ======================================================
# 5️⃣ Generate Results (per division)
# ======================================================
@admin_bp.route("/results/generate", methods=["POST"])
@token_required
@admin_required
def generate_results(user_id=None, user_type=None):
    """
    Generate / update results for a division (admin only)
    """
    division = request.json.get("division")
    if not division:
        return {"error": "division is required"}, 400

    generate_results_for_division(division)

    return {"message": f"Results generated for division {division}"}, 200


# ======================================================
# 6️⃣ Get available divisions (admin)
# ======================================================
@admin_bp.route("/divisions", methods=["GET"])
@token_required
@admin_required
def list_divisions(user_id=None, user_type=None):
    divisions = db.session.query(Student.division).distinct().all()
    divs = [d[0] for d in divisions]
    return jsonify(divs), 200


# ======================================================
# 7️⃣ Fetch results by division or roll_no (admin)
# Query params: division OR roll_no (+ optional division)
# ======================================================
@admin_bp.route("/results", methods=["GET"])
@token_required
@admin_required
def fetch_results(user_id=None, user_type=None):
    roll_no = request.args.get("roll_no")
    division = request.args.get("division")

    # If roll_no provided, optionally restrict by division
    if roll_no:
        students = Student.query.filter_by(roll_no=roll_no)
        if division:
            students = students.filter_by(division=division)
        students = students.all()
        if not students:
            return {"error": "Student not found"}, 404

        # Ensure results are generated for involved divisions
        for s in students:
            try:
                generate_results_for_division(s.division)
            except Exception:
                pass

        # Build rows for each matching student (usually one)
        rows = []
        # Prefer master Excel file data if available
        from config import MASTER_EXCEL_PATH, MASTER_EXCEL_SHEET
        use_excel = False
        try:
            import os
            if os.path.exists(MASTER_EXCEL_PATH):
                use_excel = True
        except Exception:
            use_excel = False
        for s in students:
            # If a master Excel exists, try to read detailed marks from it for this roll
            excel_marks = None
            if use_excel:
                try:
                    import openpyxl, io
                    data = openpyxl.load_workbook(MASTER_EXCEL_PATH, data_only=True)
                    if MASTER_EXCEL_SHEET in data.sheetnames:
                        sh = data[MASTER_EXCEL_SHEET]
                        # build header map
                        it = sh.iter_rows(values_only=True)
                        headers = [str(x).strip().lower() if x is not None else '' for x in next(it)]
                        def idx_of(names):
                            for n in names:
                                if n in headers:
                                    return headers.index(n)
                            return None
                        r_idx = idx_of(['roll_no', 'roll', 'rollno'])
                        d_idx = idx_of(['division', 'div'])
                        # subject column may be subject code
                        subj_idx = idx_of(['subject', 'subject_code', 'subject_id'])
                        u1_idx = idx_of(['unit1'])
                        u2_idx = idx_of(['unit2'])
                        term_idx = idx_of(['term'])
                        annual_idx = idx_of(['annual'])
                        grace_idx = idx_of(['grace'])
                        for row in it:
                            if not row or all(c is None for c in row):
                                continue
                            rv = row[r_idx] if r_idx is not None and r_idx < len(row) else None
                            dv = row[d_idx] if d_idx is not None and d_idx < len(row) else None
                            if rv is None:
                                continue
                            if str(rv).strip() == str(s.roll_no).strip() and (not division or (dv and str(dv).strip() == str(s.division).strip())):
                                # found matching excel row
                                excel_marks = {
                                    'unit1': row[u1_idx] if u1_idx is not None and u1_idx < len(row) else None,
                                    'unit2': row[u2_idx] if u2_idx is not None and u2_idx < len(row) else None,
                                    'term': row[term_idx] if term_idx is not None and term_idx < len(row) else None,
                                    'annual': row[annual_idx] if annual_idx is not None and annual_idx < len(row) else None,
                                    'grace': row[grace_idx] if grace_idx is not None and grace_idx < len(row) else None,
                                }
                                break
                except Exception:
                    excel_marks = None

            result = Result.query.filter_by(roll_no=s.roll_no, division=s.division).first()
            # If result is missing, fall back to available Marks so UI can show partial data
            marks = Mark.query.filter_by(roll_no=s.roll_no, division=s.division).all()
            subjects_map = {sub.subject_id: sub.subject_code for sub in Subject.query.all()}
            mark_map = {}
            for m in marks:
                code = subjects_map.get(m.subject_id)
                if code:
                    mark_map[code] = m

            subject_entries = []
            total_avg = 0
            total_grace = 0

            for code, field in {"ENG": "eng", "ECO": "eco", "BK": "bk", "OC": "oc"}.items():
                if result:
                    avg = getattr(result, f"{field}_avg", None)
                    grace = getattr(result, f"{field}_grace", 0) or 0
                else:
                    m = mark_map.get(code)
                    avg = m.annual if m and m.annual is not None else None
                    grace = m.grace if m and m.grace is not None else 0

                final = None
                if avg is not None:
                    final = (avg or 0) + (grace or 0)
                    total_avg += avg or 0
                    total_grace += grace or 0

                # include detailed mark breakdown if available; prefer excel row values when present
                m = mark_map.get(code)
                mark_detail = None
                if excel_marks is not None:
                    mark_detail = {
                        "unit1": excel_marks.get('unit1'),
                        "unit2": excel_marks.get('unit2'),
                        "term": excel_marks.get('term'),
                        "annual": excel_marks.get('annual'),
                        "grace": excel_marks.get('grace'),
                    }
                elif m:
                    mark_detail = {
                        "mark_id": m.mark_id,
                        "unit1": m.unit1,
                        "unit2": m.unit2,
                        "term": m.term,
                        "annual": m.annual,
                        "tot": m.tot,
                        "sub_avg": m.sub_avg,
                        "grace": m.grace,
                    }

                subject_entries.append({
                    "code": code,
                    "avg": avg,
                    "grace": grace,
                    "final": final,
                    "mark": mark_detail
                })

            # EVS and PE (grade-only) — prefer grades from Result, fall back to marks
            if result and getattr(result, 'evs_grade', None) is not None:
                subject_entries.append({"code": "EVS", "grade": result.evs_grade})
            else:
                m = mark_map.get('EVS')
                if m and m.annual is not None:
                    a = m.annual
                    grade = None
                    if a >= 75:
                        grade = 'A+'
                    elif a >= 60:
                        grade = 'A'
                    elif a >= 50:
                        grade = 'B'
                    elif a >= 35:
                        grade = 'C'
                    else:
                        grade = 'F'
                    subject_entries.append({"code": "EVS", "grade": grade, "mark": {"annual": m.annual, "mark_id": m.mark_id, "unit1": m.unit1, "unit2": m.unit2, "term": m.term, "tot": m.tot, "sub_avg": m.sub_avg, "grace": m.grace}})

            if result and getattr(result, 'pe_grade', None) is not None:
                subject_entries.append({"code": "PE", "grade": result.pe_grade})
            else:
                m = mark_map.get('PE')
                if m and m.annual is not None:
                    a = m.annual
                    grade = None
                    if a >= 75:
                        grade = 'A+'
                    elif a >= 60:
                        grade = 'A'
                    elif a >= 50:
                        grade = 'B'
                    elif a >= 35:
                        grade = 'C'
                    else:
                        grade = 'F'
                    subject_entries.append({"code": "PE", "grade": grade, "mark": {"annual": m.annual, "mark_id": m.mark_id, "unit1": m.unit1, "unit2": m.unit2, "term": m.term, "tot": m.tot, "sub_avg": m.sub_avg, "grace": m.grace}})

            for code, field in {"HINDI": "hindi", "IT": "it", "MATHS": "maths", "SP": "sp"}.items():
                include = False
                if code in ("HINDI", "IT") and s.optional_subject == code:
                    include = True
                if code in ("MATHS", "SP") and s.optional_subject_2 == code:
                    include = True
                if include:
                    avg = getattr(result, f"{field}_avg", None) if result else None
                    grace = getattr(result, f"{field}_grace", 0) if result else 0
                    final = None
                    if avg is not None:
                        final = (avg or 0) + (grace or 0)
                        total_avg += avg or 0
                        total_grace += grace or 0

                    # include detailed mark breakdown if available
                    m = mark_map.get(code)
                    mark_detail = None
                    if m:
                        mark_detail = {
                            "mark_id": m.mark_id,
                            "unit1": m.unit1,
                            "unit2": m.unit2,
                            "term": m.term,
                            "annual": m.annual,
                            "tot": m.tot,
                            "sub_avg": m.sub_avg,
                            "grace": m.grace,
                        }

                    subject_entries.append({
                        "code": code,
                        "avg": avg,
                        "grace": grace,
                        "final": final,
                        "mark": mark_detail
                    })

            final_total = None
            if subject_entries:
                # Only show final_total if percentage exists (i.e., result fully computed)
                perc = getattr(result, "percentage", None) if result else None
                if perc is not None:
                    final_total = total_avg + total_grace

            rows.append({
                "roll_no": s.roll_no,
                "name": s.name,
                "division": s.division,
                "subjects": subject_entries,
                "total_avg": round(total_avg, 2),
                "total_grace": round(total_grace, 2),
                "final_total": round(final_total, 2) if final_total is not None else None,
                "percentage": getattr(result, "percentage", None) if result else None,
            })

        # If caller requested a single roll_no, return single object
        if len(rows) == 1:
            return jsonify(rows[0]), 200
        return jsonify(rows), 200

    # Else, require division
    if not division:
        return {"error": "division or roll_no is required"}, 400

    # regenerate results for the division
    try:
        generate_results_for_division(division)
    except Exception:
        pass

    # Build rows for entire division
    students = Student.query.filter_by(division=division).order_by(Student.roll_no).all()
    rows = []
    for idx, s in enumerate(students, start=1):
        result = Result.query.filter_by(roll_no=s.roll_no, division=s.division).first()
        # Prepare mark map to allow partial display when Result row missing
        marks = Mark.query.filter_by(roll_no=s.roll_no, division=s.division).all()
        subjects_map = {sub.subject_id: sub.subject_code for sub in Subject.query.all()}
        mark_map = {}
        for m in marks:
            code = subjects_map.get(m.subject_id)
            if code:
                mark_map[code] = m

        subject_entries = []
        total_avg = 0
        total_grace = 0

        for code, field in {"ENG": "eng", "ECO": "eco", "BK": "bk", "OC": "oc"}.items():
            if result:
                avg = getattr(result, f"{field}_avg", None)
                grace = getattr(result, f"{field}_grace", 0) or 0
            else:
                m = mark_map.get(code)
                avg = m.annual if m and m.annual is not None else None
                grace = m.grace if m and m.grace is not None else 0

            final = None
            if avg is not None:
                final = (avg or 0) + (grace or 0)
                total_avg += avg or 0
                total_grace += grace or 0

            # include mark detail when present
            m = mark_map.get(code)
            mark_detail = None
            if m:
                mark_detail = {
                    "mark_id": m.mark_id,
                    "unit1": m.unit1,
                    "unit2": m.unit2,
                    "term": m.term,
                    "annual": m.annual,
                    "tot": m.tot,
                    "sub_avg": m.sub_avg,
                    "grace": m.grace,
                }

            subject_entries.append({"code": code, "avg": avg, "grace": grace, "final": final, "mark": mark_detail})

        for code, field in {"HINDI": "hindi", "IT": "it", "MATHS": "maths", "SP": "sp"}.items():
            include = False
            if code in ("HINDI", "IT") and s.optional_subject == code:
                include = True
            if code in ("MATHS", "SP") and s.optional_subject_2 == code:
                include = True

            if include:
                if result:
                    avg = getattr(result, f"{field}_avg", None)
                    grace = getattr(result, f"{field}_grace", 0) or 0
                else:
                    m = mark_map.get(code)
                    avg = m.annual if m and m.annual is not None else None
                    grace = m.grace if m and m.grace is not None else 0
                final = None
                if avg is not None:
                    final = (avg or 0) + (grace or 0)
                    total_avg += avg or 0
                    total_grace += grace or 0

                subject_entries.append({"code": code, "avg": avg, "grace": grace, "final": final})
        # EVS and PE (grade-only) — append once after optional subjects
        if result and getattr(result, 'evs_grade', None) is not None:
            subject_entries.append({"code": "EVS", "grade": result.evs_grade})
        else:
            m = mark_map.get('EVS')
            if m and m.annual is not None:
                a = m.annual
                grade = None
                if a >= 75:
                    grade = 'A+'
                elif a >= 60:
                    grade = 'A'
                elif a >= 50:
                    grade = 'B'
                elif a >= 35:
                    grade = 'C'
                else:
                    grade = 'F'
                subject_entries.append({"code": "EVS", "grade": grade, "mark": {"annual": m.annual, "mark_id": m.mark_id, "unit1": m.unit1, "unit2": m.unit2, "term": m.term, "tot": m.tot, "sub_avg": m.sub_avg, "grace": m.grace}})

        if result and getattr(result, 'pe_grade', None) is not None:
            subject_entries.append({"code": "PE", "grade": result.pe_grade})
        else:
            m = mark_map.get('PE')
            if m and m.annual is not None:
                a = m.annual
                grade = None
                if a >= 75:
                    grade = 'A+'
                elif a >= 60:
                    grade = 'A'
                elif a >= 50:
                    grade = 'B'
                elif a >= 35:
                    grade = 'C'
                else:
                    grade = 'F'
                subject_entries.append({"code": "PE", "grade": grade, "mark": {"annual": m.annual, "mark_id": m.mark_id, "unit1": m.unit1, "unit2": m.unit2, "term": m.term, "tot": m.tot, "sub_avg": m.sub_avg, "grace": m.grace}})

        final_total = None
        if subject_entries:
            perc = getattr(result, "percentage", None) if result else None
            if perc is not None:
                final_total = total_avg + total_grace

        rows.append({
            "seq": idx,
            "roll_no": s.roll_no,
            "name": s.name,
            "subjects": subject_entries,
            "total_avg": round(total_avg, 2),
            "total_grace": round(total_grace, 2),
            "final_total": round(final_total, 2) if final_total is not None else None,
            "percentage": getattr(result, "percentage", None) if result else None,
        })

    return jsonify(rows), 200


@admin_bp.route('/excel/master', methods=['GET'])
@token_required
@admin_required
def download_master_excel(user_id=None, user_type=None):
    """Allow admin to download or open the shared master Excel file."""
    from config import MASTER_EXCEL_PATH
    import os
    if not os.path.exists(MASTER_EXCEL_PATH):
        return {"error": "Master Excel file not found on server"}, 404
    try:
        return send_file(MASTER_EXCEL_PATH, as_attachment=True)
    except Exception as ex:
        return {"error": "Failed to send master Excel file", "details": str(ex)}, 500


@admin_bp.route('/excel/student', methods=['GET'])
@token_required
@admin_required
def download_student_excel(user_id=None, user_type=None):
    """Generate an Excel file for a single student in the master format and return it."""
    roll_no = request.args.get('roll_no')
    division = request.args.get('division')
    if not roll_no or not division:
        return {"error": "roll_no and division are required"}, 400

    # Fetch student and marks
    student = Student.query.filter_by(roll_no=roll_no, division=division).first()
    if not student:
        return {"error": "Student not found"}, 404

    # Prepare per-subject rows: iterate through all subjects student takes
    all_subjects = Subject.query.filter_by(active=True).order_by(Subject.subject_code).all()
    include_codes = set()
    for s in all_subjects:
        if s.subject_type == 'CORE':
            include_codes.add(s.subject_code)
    if student.optional_subject:
        include_codes.add(student.optional_subject)
    if student.optional_subject_2:
        include_codes.add(student.optional_subject_2)

    subjects = [s for s in all_subjects if s.subject_code in include_codes]

    # Build rows
    rows = []
    for s in subjects:
        m = Mark.query.filter_by(roll_no=roll_no, division=division, subject_id=s.subject_id).first()
        teacher_name = None
        if m and m.entered_by:
            t = Teacher.query.get(m.entered_by)
            teacher_name = t.name if t else None

        rows.append({
            'roll_no': roll_no,
            'student_name': student.name,
            'subject': s.subject_code,
            'division': division,
            'unit1': m.unit1 if m else None,
            'unit2': m.unit2 if m else None,
            'term': m.term if m else None,
            'annual': m.annual if m else None,
            'tot': m.tot if m else None,
            'sub_avg': m.sub_avg if m else None,
            'grace': m.grace if m else None,
            'final': (m.sub_avg + (m.grace or 0)) if m and m.sub_avg is not None else None,
            'entered_by': teacher_name
        })

    # Create Excel
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = cast(Any, wb.active)
        ws.title = 'Marks'
        headers = ['Roll', 'Student Name', 'Subject', 'Division', 'Unit1', 'Unit2', 'Term', 'Annual', 'Tot', 'Sub_Avg', 'Grace', 'Final', 'Entered By']
        ws.append(headers)
        for r in rows:
            ws.append([
                r['roll_no'],
                r['student_name'],
                r['subject'],
                r['division'],
                r['unit1'] if r['unit1'] is not None else '',
                r['unit2'] if r['unit2'] is not None else '',
                r['term'] if r['term'] is not None else '',
                r['annual'] if r['annual'] is not None else '',
                r['tot'] if r['tot'] is not None else '',
                r['sub_avg'] if r['sub_avg'] is not None else '',
                r['grace'] if r['grace'] is not None else '',
                r['final'] if r['final'] is not None else '',
                r['entered_by'] or ''
            ])

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"student_{roll_no}_{division}.xlsx"
        return send_file(bio, download_name=filename, as_attachment=True)
    except Exception as ex:
        return {"error": "Failed to generate Excel", "details": str(ex)}, 500


@admin_bp.route('/excel/complete', methods=['GET'])
@token_required
@admin_required
def download_complete_excel(user_id=None, user_type=None):
    """Generate the complete merged result Excel in the exact required column order.

    Query params:
      - division (required unless roll_no provided)
      - roll_no (optional)
    """
    roll_no = request.args.get('roll_no')
    division = request.args.get('division')

    # Determine target students
    if roll_no:
        students_q = Student.query.filter_by(roll_no=roll_no)
        if division:
            students_q = students_q.filter_by(division=division)
        students = students_q.all()
        if not students:
            return {"error": "Student not found"}, 404
    else:
        if not division:
            return {"error": "division or roll_no is required"}, 400
        students = Student.query.filter_by(division=division).order_by(Student.roll_no).all()
        if not students:
            return {"error": "No students found for division"}, 404

    all_subjects = Subject.query.filter_by(active=True).order_by(Subject.subject_code).all()

    rows = []
    # Ensure results are generated for each involved division
    try:
        divisions = set(s.division for s in students)
        for d in divisions:
            try:
                generate_results_for_division(d)
            except Exception:
                pass
    except Exception:
        pass
    for s in students:
        include_codes = set()
        for sub in all_subjects:
            if sub.subject_type == 'CORE':
                include_codes.add(sub.subject_code)
        if s.optional_subject:
            include_codes.add(s.optional_subject)
        if s.optional_subject_2:
            include_codes.add(s.optional_subject_2)

        for code in sorted(include_codes):
            subj = next((x for x in all_subjects if x.subject_code == code), None)
            if not subj:
                continue
            m = Mark.query.filter_by(roll_no=s.roll_no, division=s.division, subject_id=subj.subject_id).first()

            rows.append({
                'Roll': s.roll_no,
                'Student Name': s.name,
                'Subject': code,
                'Division': s.division,
                'Unit1': m.unit1 if m and m.unit1 is not None else '',
                'Term': m.term if m and m.term is not None else '',
                'Unit2': m.unit2 if m and m.unit2 is not None else '',
                'Annual': m.annual if m and m.annual is not None else '',
                'Grace': m.grace if m and m.grace is not None else ''
            })

    headers = [
        "Roll",
        "Student Name",
        "Subject",
        "Division",
        "Unit1",
        "Term",
        "Unit2",
        "Annual",
        "Grace",
    ]

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = cast(Any, wb.active)
        ws.title = 'Complete Results'
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, '') for h in headers])

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        fn = 'complete_results'
        if division:
            fn += f'_{division}'
        if roll_no:
            fn += f'_roll_{roll_no}'
        fn += '.xlsx'
        return send_file(bio, download_name=fn, as_attachment=True)
    except Exception as ex:
        return {"error": "Failed to generate Excel", "details": str(ex)}, 500


@admin_bp.route('/excel/division', methods=['GET'])
@token_required
@admin_required
def download_division_excel(user_id=None, user_type=None):
    division = request.args.get('division')
    if not division:
        return {"error": "division is required"}, 400

    # Build rows for all students in division
    students = Student.query.filter_by(division=division).order_by(Student.roll_no).all()
    if not students:
        return {"error": "No students found for division"}, 404

    rows = []
    all_subjects = Subject.query.filter_by(active=True).order_by(Subject.subject_code).all()
    subjects_codes = [s.subject_code for s in all_subjects]

    for s in students:
        include_codes = set()
        for sub in all_subjects:
            if sub.subject_type == 'CORE':
                include_codes.add(sub.subject_code)
        if s.optional_subject:
            include_codes.add(s.optional_subject)
        if s.optional_subject_2:
            include_codes.add(s.optional_subject_2)

        for code in sorted(include_codes):
            subj = next((x for x in all_subjects if x.subject_code == code), None)
            if not subj:
                continue
            m = Mark.query.filter_by(roll_no=s.roll_no, division=s.division, subject_id=subj.subject_id).first()
            teacher_name = None
            if m and m.entered_by:
                t = Teacher.query.get(m.entered_by)
                teacher_name = t.name if t else None
            rows.append({
                'roll_no': s.roll_no,
                'student_name': s.name,
                'subject': code,
                'division': s.division,
                'unit1': m.unit1 if m else None,
                'unit2': m.unit2 if m else None,
                'term': m.term if m else None,
                'annual': m.annual if m else None,
                'grace': m.grace if m else None,
                'entered_by': teacher_name
            })

    # Create Excel file
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = cast(Any, wb.active)
        ws.title = 'Marks'
        headers = ['Roll', 'Student Name', 'Subject', 'Division', 'Unit1', 'Unit2', 'Term', 'Annual', 'Grace', 'Entered By']
        ws.append(headers)
        for r in rows:
            ws.append([
                r['roll_no'], r['student_name'], r['subject'], r['division'],
                r['unit1'] if r['unit1'] is not None else '',
                r['unit2'] if r['unit2'] is not None else '',
                r['term'] if r['term'] is not None else '',
                r['annual'] if r['annual'] is not None else '',
                r['grace'] if r['grace'] is not None else '',
                r['entered_by'] or ''
            ])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"division_{division}_marks.xlsx"
        return send_file(bio, download_name=filename, as_attachment=True)
    except Exception as ex:
        return {"error": "Failed to generate Excel", "details": str(ex)}, 500


@admin_bp.route('/excel/marksheet', methods=['GET'])
@token_required
@admin_required
def download_marksheet(user_id=None, user_type=None):
    """Generate final printable marksheet Excel for a division.

    Query params:
      - division (required)
    The sheet follows the strict layout required by admin (merged headers,
    fixed subject blocks and internal columns, max marks row, one student per row).
    """
    division = request.args.get('division')
    if not division:
        return {"error": "division is required"}, 400

    # Ensure computed results exist
    try:
        generate_results_for_division(division)
    except Exception:
        pass

    # Load students for division
    students = Student.query.filter_by(division=division).order_by(Student.roll_no).all()
    if not students:
        return {"error": "No students found for division"}, 404

    # Fixed subject order required by spec
    SUBJECT_ORDER = [
        ('ENGLISH', ['ENG']),
        ('OC', ['OC']),
        ('SP / MATHS', ['SP', 'MATHS']),
        ('ECONOMICS', ['ECO']),
        ('B.K. & A/C', ['BK']),
    ]

    # Map subject_code -> Subject object
    all_subjects = Subject.query.filter_by(active=True).all()
    subjects_by_code = {s.subject_code.upper(): s for s in all_subjects}

    def resolve_optional_subject(student, candidates):
        # Check student's optional_subject then optional_subject_2
        for opt in (student.optional_subject, student.optional_subject_2):
            if opt and opt.upper() in candidates:
                return subjects_by_code.get(opt.upper())
        # fallback: any available candidate in DB
        for c in candidates:
            if c.upper() in subjects_by_code:
                return subjects_by_code[c.upper()]
        return None

    # Build workbook
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        wb = Workbook()
        ws = cast(Any, wb.active)
        ws.title = 'Marksheet'

        # columns: 2 ID columns + 5 subjects * 8 internal cols
        total_cols = 2 + len(SUBJECT_ORDER) * 8

        # Top headers (merged)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws.cell(row=1, column=1, value='SIES COLLEGE OF COMMERCE, NERUL')
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        title = f'FYJC (DIV {division}) MARKSHEET – 2024–2025'
        ws.cell(row=2, column=1, value=title)
        ws.cell(row=2, column=1).font = Font(bold=True)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

        # Row 3 right-aligned stream text
        ws.cell(row=3, column=total_cols, value='B.K. & A/C')
        ws.cell(row=3, column=total_cols).alignment = Alignment(horizontal='right')

        # Header rows: row4 merged subject headers, row5 internal headers
        ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
        ws.cell(row=4, column=1, value='ROLL NO')
        ws.cell(row=4, column=1).font = Font(bold=True)
        ws.cell(row=4, column=1).alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)
        ws.cell(row=4, column=2, value='STUDENT NAME')
        ws.cell(row=4, column=2).font = Font(bold=True)
        ws.cell(row=4, column=2).alignment = Alignment(horizontal='center', vertical='center')

        internal_headers = ['UNIT I', 'TERM I', 'UNIT II', 'INT', 'ANNUAL', 'TOT', 'AVG', 'GRACE']
        col = 3
        for subj_label, candidates in SUBJECT_ORDER:
            start = col
            end = col + len(internal_headers) - 1
            ws.merge_cells(start_row=4, start_column=start, end_row=4, end_column=end)
            ws.cell(row=4, column=start, value=subj_label)
            ws.cell(row=4, column=start).font = Font(bold=True)
            ws.cell(row=4, column=start).alignment = Alignment(horizontal='center')
            for i, h in enumerate(internal_headers):
                c = ws.cell(row=5, column=start + i, value=h)
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal='center')
            col += len(internal_headers)

        # Row 6: Maximum marks
        max_row = 6
        ws.cell(row=max_row, column=1, value='')
        ws.cell(row=max_row, column=2, value='')
        max_values = [25, 50, 25, 20, 80, 200, 100, '']
        col = 3
        for _ in SUBJECT_ORDER:
            for mv in max_values:
                ws.cell(row=max_row, column=col, value=mv)
                ws.cell(row=max_row, column=col).alignment = Alignment(horizontal='center')
                col += 1

        # Student rows
        row_idx = 7
        for s in students:
            ws.cell(row=row_idx, column=1, value=s.roll_no)
            ws.cell(row=row_idx, column=2, value=s.name)
            col = 3
            for subj_label, candidates in SUBJECT_ORDER:
                # resolve subject
                if subj_label == 'SP / MATHS':
                    subj_obj = resolve_optional_subject(s, [c.upper() for c in candidates])
                else:
                    subj_obj = None
                    for code in candidates:
                        if code.upper() in subjects_by_code:
                            subj_obj = subjects_by_code[code.upper()]
                            break

                if not subj_obj:
                    # leave 8 blank cells
                    for i in range(8):
                        ws.cell(row=row_idx, column=col + i, value='')
                    col += 8
                    continue

                m = Mark.query.filter_by(roll_no=s.roll_no, division=s.division, subject_id=subj_obj.subject_id).first()
                if m:
                    unit1 = m.unit1 if m.unit1 is not None else ''
                    term = m.term if m.term is not None else ''
                    unit2 = m.unit2 if m.unit2 is not None else ''
                    internal = m.internal if m.internal is not None else ''
                    annual = m.annual if m.annual is not None else ''
                    # TOT and AVG per rules
                    if m.tot is not None and m.tot != 0:
                        tot = m.tot
                    else:
                        try:
                            tot = (m.unit1 or 0) + (m.term or 0) + (m.unit2 or 0) + (m.internal or 0) + (m.annual or 0)
                        except Exception:
                            tot = ''
                    if m.sub_avg is not None and m.sub_avg != 0:
                        avg = m.sub_avg
                    else:
                        try:
                            avg = round(float(tot) / 2) if tot != '' else ''
                        except Exception:
                            avg = ''
                    grace = m.grace if m.grace is not None else ''
                else:
                    unit1 = term = unit2 = internal = annual = tot = avg = grace = ''

                values = [unit1, term, unit2, internal, annual, tot, avg, grace]
                for i, v in enumerate(values):
                    cell = ws.cell(row=row_idx, column=col + i, value=v)
                    cell.alignment = Alignment(horizontal='center')
                col += 8

            row_idx += 1

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f'marksheet_div_{division}.xlsx'
        return send_file(bio, download_name=filename, as_attachment=True)
    except Exception as ex:
        return {"error": "Failed to generate marksheet", "details": str(ex)}, 500


# ======================================================
# Download student marksheet PDF (admin only)
# ======================================================
@admin_bp.route('/students/<string:roll_no>/pdf', methods=['GET'])
@token_required
@admin_required
def student_marksheet_pdf(roll_no, user_id=None, user_type=None):
    division = request.args.get('division')
    if not division:
        return {"error": "division is required"}, 400

    # ensure results are up-to-date
    try:
        generate_results_for_division(division)
    except Exception:
        pass

    res = Result.query.filter_by(roll_no=roll_no, division=division).first()
    if not res:
        return {"error": "Result not found"}, 404

    if canvas_module is None or letter is None:
        return {"error": "reportlab not installed on server. Install reportlab in requirements."}, 501

    buf = BytesIO()
    # canvas_module is the imported reportlab.pdfgen.canvas module
    CanvasClass = getattr(canvas_module, 'Canvas', None)
    if CanvasClass is None:
        return {"error": "reportlab canvas API not available"}, 501
    c = CanvasClass(buf, pagesize=cast(Any, letter))
    width, height = cast(Any, letter)

    # Header
    c.setFont('Helvetica-Bold', 16)
    c.drawString(40, height - 50, 'Official Marksheet')
    c.setFont('Helvetica', 12)
    c.drawString(40, height - 70, f'Name: {res.name}  |  Roll: {res.roll_no}  |  Division: {res.division}')

    # Table header
    y = height - 110
    c.setFont('Helvetica-Bold', 11)
    c.drawString(40, y, 'Subject')
    c.drawString(260, y, 'Annual')
    c.drawString(360, y, 'Grace')
    c.drawString(460, y, 'Final')
    c.setFont('Helvetica', 11)
    y -= 18

    # iterate known subjects and map to result fields
    mapping = [
        ('ENG', 'eng_avg', 'eng_grace'),
        ('ECO', 'eco_avg', 'eco_grace'),
        ('BK', 'bk_avg', 'bk_grace'),
        ('OC', 'oc_avg', 'oc_grace'),
        ('HINDI', 'hindi_avg', 'hindi_grace'),
        ('IT', 'it_avg', 'it_grace'),
        ('MATHS', 'maths_avg', 'maths_grace'),
        ('SP', 'sp_avg', 'sp_grace')
    ]

    total = 0
    for code, avg_field, grace_field in mapping:
        avg = getattr(res, avg_field, None)
        grace = getattr(res, grace_field, 0) or 0
        if avg is None:
            continue
        final = (avg or 0) + (grace or 0)
        c.drawString(40, y, code)
        c.drawRightString(320, y, f'{round(avg,2)}')
        c.drawRightString(420, y, f'{round(grace,2)}')
        c.drawRightString(520, y, f'{round(final,2)}')
        total += final
        y -= 16

    y -= 8
    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, y, f'Total: {round(res.percentage * (len(mapping) / len(mapping)),2) if res.percentage is not None else "-"}')
    c.drawRightString(520, y, f'Percentage: {res.percentage or "-"}')

    c.showPage()
    c.save()
    buf.seek(0)

    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=f'{roll_no}_marksheet.pdf')



# ======================================================
# ADMIN – TEACHERS CRUD
# ======================================================

@admin_bp.route("/teachers", methods=["GET"])
@token_required
def list_teachers(user_id=None, user_type=None):
    if user_type != "ADMIN":
        return {"error": "Unauthorized"}, 403

    teachers = Teacher.query.all()
    return jsonify([
        {
            "teacher_id": t.teacher_id,
            "name": t.name,
            "userid": t.userid,
            "email": t.email,
            "active": t.active,
            "role": t.role
        }
        for t in teachers
    ]), 200


@admin_bp.route("/teachers", methods=["POST"])
@token_required
def add_teacher(user_id=None, user_type=None):
    if user_type != "ADMIN":
        return {"error": "Unauthorized"}, 403

    data = request.json
    required = ["name", "userid", "password"]
    if not all(k in data for k in required):
        return {"error": "Missing required fields"}, 400

    if Teacher.query.filter_by(userid=data["userid"]).first():
        return {"error": "UserID already exists"}, 409

    teacher = Teacher(
        name=data["name"],
        userid=data["userid"],
        email=data.get("email"),
        role=data.get("role", "TEACHER"),
        active=True,
        password_hash=generate_password_hash(data["password"])
    )

    db.session.add(teacher)
    db.session.commit()

    return {"message": "Teacher added"}, 201


@admin_bp.route("/teachers/<int:teacher_id>", methods=["PUT"])
@token_required
def update_teacher(teacher_id, user_id=None, user_type=None):
    if user_type != "ADMIN":
        return {"error": "Unauthorized"}, 403

    teacher = Teacher.query.get_or_404(teacher_id)
    data = request.json

    teacher.name = data.get("name", teacher.name)
    teacher.userid = data.get("userid", teacher.userid)
    teacher.email = data.get("email", teacher.email)
    teacher.active = data.get("active", teacher.active)

    if data.get("password"):
        teacher.password_hash = generate_password_hash(data["password"])

    db.session.commit()
    return {"message": "Teacher updated"}, 200


@admin_bp.route("/teachers/<int:teacher_id>", methods=["DELETE"])
@token_required
def delete_teacher(teacher_id, user_id=None, user_type=None):
    if user_type != "ADMIN":
        return {"error": "Unauthorized"}, 403

    teacher = Teacher.query.get_or_404(teacher_id)
    db.session.delete(teacher)
    db.session.commit()

    return {"message": "Teacher deleted"}, 200


# ======================================================
# ADMIN LOGIN (tests expect /admin/login)
# ======================================================
@admin_bp.route("/login", methods=["POST"])
def admin_login():
    data = request.json or {}
    userid = data.get("userid")
    password = data.get("password")

    if not userid or not password:
        return {"error": "userid and password required"}, 400

    admin = None
    from models import Admin
    admin = Admin.query.filter_by(username=userid, active=True).first()
    if not admin:
        return {"error": "Invalid credentials"}, 401

    # Admin passwords are hashed with werkzeug.generate_password_hash
    from werkzeug.security import check_password_hash
    if not check_password_hash(admin.password_hash, password):
        return {"error": "Invalid credentials"}, 401

    token = generate_token(admin.admin_id, "ADMIN")
    # Return role in lowercase for consistency with client/tests
    return {"token": token, "role": "admin"}, 200


# ======================================================
# ADMIN IMPERSONATE TEACHER (open teacher panel without logging out)
# ======================================================
@admin_bp.route('/teachers/<int:teacher_id>/impersonate', methods=['POST'])
@token_required
@admin_required
def impersonate_teacher(teacher_id, user_id=None, user_type=None):
    """
    Generate a short-lived token for a teacher so an admin can open
    the teacher panel without logging out (impersonation).
    """
    teacher = Teacher.query.get(teacher_id)
    if not teacher or not teacher.active:
        return {"error": "Teacher not found or inactive"}, 404

    # Issue token with TEACHER role
    token = generate_token(teacher.teacher_id, "TEACHER", expires_hours=2)

    return {
        "token": token,
        "teacher": {
            "teacher_id": teacher.teacher_id,
            "name": teacher.name,
            "userid": teacher.userid,
            "email": teacher.email,
        }
    }, 200
